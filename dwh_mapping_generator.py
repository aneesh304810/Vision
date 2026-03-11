"""
DWH Mapping Generator
Reads a 3-sheet Excel workbook (SRC→STG1, STG1→STG2, STG2→DWH)
and produces a consolidated Data Warehouse Mapping Excel file.

Usage:
    python dwh_mapping_generator.py --input mappings.xlsx --output dwh_mapping.xlsx
"""

import argparse
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXPECTED_SHEETS = ["SRC TO STG1", "STG1 TO STG2", "STG2 TO DWH"]

INPUT_COLUMNS = [
    "category",
    "field description",
    "source table name",
    "sources column name",
    "targets database table - name",
    "targets database column name",
    "targets database column - native type",
    "target application transformation rule",
]

OUTPUT_COLUMNS = [
    "Field",
    "Table",
    "Field Description",
    "Source (SRC)",
    "SRC Field Name",
    "SRC Field Type",
    "SRC Table Name",
    "Load Type",
    "Staging 1 - Table",
    "Staging 1 - Field",
    "Staging 1 - Transformation Type",
    "Staging 1 - Transformation Logic",
    "Staging 2 - Table",
    "Staging 2 - Field",
    "Staging 2 - Transformation Type",
    "Staging 2 - Transformation Logic",
    "Target - Table",
    "Target - Field",
    "Target - Transformation Type",
    "Target - Transformation Logic",
    "Notes",
    "Purpose of Job / Description",
]

HEADER_FILL_COLOR = "1F4E79"   # dark blue
HEADER_FONT_COLOR = "FFFFFF"   # white
ALT_ROW_COLOR     = "D6E4F0"   # light blue for alternating rows


# ---------------------------------------------------------------------------
# Step 1 – Read Input Excel
# ---------------------------------------------------------------------------

def read_input_excel(filepath: str) -> dict[str, pd.DataFrame]:
    """
    Read the three mapping sheets from the input workbook.
    Returns a dict keyed by normalised sheet name.
    Raises ValueError if expected sheets are missing.
    """
    try:
        xl = pd.ExcelFile(filepath)
    except FileNotFoundError:
        sys.exit(f"[ERROR] Input file not found: {filepath}")

    # Normalise available sheet names for case-insensitive matching
    available = {s.strip().upper(): s for s in xl.sheet_names}
    sheets: dict[str, pd.DataFrame] = {}

    for expected in EXPECTED_SHEETS:
        key = expected.upper()
        if key not in available:
            sys.exit(
                f"[ERROR] Expected sheet '{expected}' not found. "
                f"Available sheets: {xl.sheet_names}"
            )
        df = xl.parse(available[key])
        df.columns = [str(c).strip().lower() for c in df.columns]   # normalise headers
        df = df.dropna(how="all")                                    # drop blank rows
        sheets[expected] = df

    print(f"[INFO] Read sheets: {list(sheets.keys())}")
    for name, df in sheets.items():
        print(f"       {name}: {len(df)} rows, columns: {list(df.columns)}")

    return sheets


# ---------------------------------------------------------------------------
# Step 2 – Build per-layer mapping DataFrames
# ---------------------------------------------------------------------------

def _get_col(df: pd.DataFrame, normalised_name: str, sheet_name: str) -> pd.Series:
    """Return a column by its normalised name; raise a clear error if missing."""
    if normalised_name not in df.columns:
        raise ValueError(
            f"Column '{normalised_name}' not found in sheet '{sheet_name}'. "
            f"Available: {list(df.columns)}"
        )
    return df[normalised_name]


def _transformation_type(rule_series: pd.Series) -> pd.Series:
    """Return 'Direct Mapping' for N/A rules, 'Derived' otherwise."""
    return rule_series.apply(
        lambda v: "Direct Mapping"
        if pd.isna(v) or str(v).strip().upper() in ("N/A", "NA", "NONE", "")
        else "Derived"
    )


def _transformation_logic(rule_series: pd.Series) -> pd.Series:
    """Return None for N/A rules, the rule string otherwise."""
    return rule_series.apply(
        lambda v: None
        if pd.isna(v) or str(v).strip().upper() in ("N/A", "NA", "NONE", "")
        else str(v).strip()
    )


def build_mapping(sheets: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    """
    Extract clean, consistently-named DataFrames from each sheet.
    Returns a dict with keys 'src_stg1', 'stg1_stg2', 'stg2_dwh'.
    """
    # ---- SRC → STG1 --------------------------------------------------------
    s1 = sheets["SRC TO STG1"].copy()
    src_stg1 = pd.DataFrame({
        "category":           _get_col(s1, "category",                                 "SRC TO STG1"),
        "field_description":  _get_col(s1, "field description",                        "SRC TO STG1"),
        "src_table":          _get_col(s1, "source table name",                        "SRC TO STG1"),
        "src_field":          _get_col(s1, "sources column name",                      "SRC TO STG1"),
        "src_field_type":     _get_col(s1, "targets database column - native type",    "SRC TO STG1"),
        "stg1_table":         _get_col(s1, "targets database table - name",            "SRC TO STG1"),
        "stg1_field":         _get_col(s1, "targets database column name",             "SRC TO STG1"),
        "stg1_rule":          _get_col(s1, "target application transformation rule",   "SRC TO STG1"),
    })

    # ---- STG1 → STG2 -------------------------------------------------------
    s2 = sheets["STG1 TO STG2"].copy()
    stg1_stg2 = pd.DataFrame({
        "stg1_table":  _get_col(s2, "source table name",                       "STG1 TO STG2"),
        "stg1_field":  _get_col(s2, "sources column name",                     "STG1 TO STG2"),
        "stg2_table":  _get_col(s2, "targets database table - name",           "STG1 TO STG2"),
        "stg2_field":  _get_col(s2, "targets database column name",            "STG1 TO STG2"),
        "stg2_rule":   _get_col(s2, "target application transformation rule",  "STG1 TO STG2"),
    })

    # ---- STG2 → DWH --------------------------------------------------------
    s3 = sheets["STG2 TO DWH"].copy()
    stg2_dwh = pd.DataFrame({
        "stg2_table":   _get_col(s3, "source table name",                      "STG2 TO DWH"),
        "stg2_field":   _get_col(s3, "sources column name",                    "STG2 TO DWH"),
        "target_table": _get_col(s3, "targets database table - name",          "STG2 TO DWH"),
        "target_field": _get_col(s3, "targets database column name",           "STG2 TO DWH"),
        "target_rule":  _get_col(s3, "target application transformation rule", "STG2 TO DWH"),
    })

    # Normalise key columns (strip whitespace, uppercase) for reliable joins
    for df, cols in [
        (src_stg1,  ["stg1_table", "stg1_field"]),
        (stg1_stg2, ["stg1_table", "stg1_field", "stg2_table", "stg2_field"]),
        (stg2_dwh,  ["stg2_table", "stg2_field"]),
    ]:
        for c in cols:
            df[c] = df[c].astype(str).str.strip().str.upper()

    return {"src_stg1": src_stg1, "stg1_stg2": stg1_stg2, "stg2_dwh": stg2_dwh}


# ---------------------------------------------------------------------------
# Step 3 – Merge layers
# ---------------------------------------------------------------------------

def merge_layers(mapping: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Left-join SRC→STG1 with STG1→STG2 on (stg1_table, stg1_field),
    then left-join that result with STG2→DWH on (stg2_table, stg2_field).
    Missing joins are preserved (NaN) and flagged in console output.
    """
    src_stg1  = mapping["src_stg1"]
    stg1_stg2 = mapping["stg1_stg2"]
    stg2_dwh  = mapping["stg2_dwh"]

    # Join 1: SRC→STG1 ⟕ STG1→STG2
    merged = src_stg1.merge(
        stg1_stg2,
        on=["stg1_table", "stg1_field"],
        how="left",
        suffixes=("", "_s2"),
    )

    unmatched_stg2 = merged["stg2_table"].isna().sum()
    if unmatched_stg2:
        print(f"[WARN] {unmatched_stg2} rows in SRC→STG1 have no match in STG1→STG2 "
              f"(stg1_table/stg1_field key). They will have blank STG2 columns.")

    # Join 2: result ⟕ STG2→DWH
    # Normalise the stg2 key in the merged frame first
    for c in ["stg2_table", "stg2_field"]:
        merged[c] = merged[c].astype(str).str.strip().str.upper()

    merged = merged.merge(
        stg2_dwh,
        on=["stg2_table", "stg2_field"],
        how="left",
        suffixes=("", "_dwh"),
    )

    unmatched_target = merged["target_table"].isna().sum()
    if unmatched_target:
        print(f"[WARN] {unmatched_target} rows have no match in STG2→DWH "
              f"(stg2_table/stg2_field key). They will have blank Target columns.")

    print(f"[INFO] Merged dataframe: {len(merged)} rows")
    return merged


# ---------------------------------------------------------------------------
# Step 4 – Build final output DataFrame
# ---------------------------------------------------------------------------

def build_output_dataframe(merged: pd.DataFrame) -> pd.DataFrame:
    """
    Map the merged dataframe onto the 22-column output schema.
    """
    out = pd.DataFrame()

    # Field / Table come from the final target layer
    out["Field"]                              = merged["target_field"]
    out["Table"]                              = merged["target_table"]
    out["Field Description"]                  = merged["field_description"]

    # Source layer
    out["Source (SRC)"]                       = merged["category"]
    out["SRC Field Name"]                     = merged["src_field"]
    out["SRC Field Type"]                     = merged["src_field_type"]
    out["SRC Table Name"]                     = merged["src_table"]
    out["Load Type"]                          = "Intraday"

    # STG1
    out["Staging 1 - Table"]                  = merged["stg1_table"]
    out["Staging 1 - Field"]                  = merged["stg1_field"]
    out["Staging 1 - Transformation Type"]    = _transformation_type(merged["stg1_rule"])
    out["Staging 1 - Transformation Logic"]   = _transformation_logic(merged["stg1_rule"])

    # STG2
    out["Staging 2 - Table"]                  = merged["stg2_table"]
    out["Staging 2 - Field"]                  = merged["stg2_field"]
    out["Staging 2 - Transformation Type"]    = _transformation_type(merged["stg2_rule"])
    out["Staging 2 - Transformation Logic"]   = _transformation_logic(merged["stg2_rule"])

    # Target
    out["Target - Table"]                     = merged["target_table"]
    out["Target - Field"]                     = merged["target_field"]
    out["Target - Transformation Type"]       = _transformation_type(merged["target_rule"])
    out["Target - Transformation Logic"]      = _transformation_logic(merged["target_rule"])

    # Blanks
    out["Notes"]                              = ""
    out["Purpose of Job / Description"]       = ""

    # Replace "NAN" strings (from normalisation) with proper empty values
    out.replace("NAN", pd.NA, inplace=True)

    return out[OUTPUT_COLUMNS]


# ---------------------------------------------------------------------------
# Step 5 – Generate formatted output Excel
# ---------------------------------------------------------------------------

def generate_output_excel(df: pd.DataFrame, output_path: str) -> None:
    """
    Write the output dataframe to a formatted Excel file with:
      - Styled header row
      - Alternating row colours
      - Auto-width columns
      - Freeze panes
    """
    # Write raw data first using openpyxl engine
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="DWH Mapping")
        wb = writer.book
        ws = writer.sheets["DWH Mapping"]

        # ---- Header styling ------------------------------------------------
        header_font  = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        header_fill  = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border  = Border(
            bottom=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="AAAAAA"),
        )

        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin_border

        ws.row_dimensions[1].height = 30

        # ---- Alternating row colours ---------------------------------------
        alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

        # ---- Auto-width columns --------------------------------------------
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Cap width at 45 characters to avoid overly wide columns
            adjusted = min(max_len + 4, 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        # ---- Freeze top row ------------------------------------------------
        ws.freeze_panes = "A2"

        # ---- Auto-filter ---------------------------------------------------
        ws.auto_filter.ref = ws.dimensions

    print(f"[INFO] Output written to: {output_path}")


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def run(input_path: str, output_path: str) -> None:
    print(f"[INFO] Reading input:  {input_path}")
    sheets  = read_input_excel(input_path)
    mapping = build_mapping(sheets)
    merged  = merge_layers(mapping)
    output  = build_output_dataframe(merged)
    generate_output_excel(output, output_path)
    print("[INFO] Done.")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate a DWH Mapping Excel from a 3-sheet pipeline mapping workbook."
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Path to the input Excel file (must contain SRC TO STG1, STG1 TO STG2, STG2 TO DWH sheets)"
    )
    parser.add_argument(
        "--output", "-o",
        default="dwh_mapping_output.xlsx",
        help="Path for the generated output Excel file (default: dwh_mapping_output.xlsx)"
    )
    args = parser.parse_args()
    run(args.input, args.output)
